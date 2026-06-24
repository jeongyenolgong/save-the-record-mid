#!/usr/bin/env python3
# 이 레포의 콘텐츠 엑셀(content/*.xlsx) → content.json 변환
# 사용법:  python3 convert.py
import json, os, glob
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SECTION_ORDER = ["생산 배경","생산 방법","생존 이야기","기록문화적 가치","현재 보존 방법","접근·열람 방법"]
METHOD = {"web":"웹제공", "card":"카드"}

def cols(ws):
    return {ws.cell(1,c).value: c for c in range(1, ws.max_column+1)}

def find_xlsx():
    xs = [x for x in glob.glob(os.path.join(HERE,"content","*.xlsx"))
          if not os.path.basename(x).startswith("~$")]
    if not xs:
        raise SystemExit("content/ 폴더에 .xlsx 가 없습니다.")
    return sorted(xs)[0]

def convert(xlsx_path, json_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    it, co = wb["items"], wb["content"]
    ic, cc = cols(it), cols(co)
    meta = {}
    for r in range(2, it.max_row+1):
        no = it.cell(r, ic["no"]).value
        if no in (None,""): continue
        no = int(no)
        gi = lambda k: it.cell(r, ic[k]).value
        meta[no] = dict(
            no=no, category=gi("category") or "",
            created=int(gi("created")) if gi("created") not in (None,"") else None,
            era_label=str(gi("era_label") or ""), name=gi("name") or "",
            country="한국",
            year=int(gi("year")) if gi("year") not in (None,"") else None,
            difficulty=int(gi("difficulty")) if gi("difficulty") not in (None,"") else None,
            desc=gi("desc") or "", items={})
    for r in range(2, co.max_row+1):
        no = co.cell(r, cc["no"]).value
        if no in (None,"") or int(no) not in meta: continue
        no = int(no)
        label = co.cell(r, cc["section_label"]).value
        code = co.cell(r, cc["code"]).value
        code = "" if code in (None,"") else str(int(code) if isinstance(code,(int,float)) else code)
        meta[no]["items"][label] = dict(
            method=METHOD.get(co.cell(r, cc["method"]).value, "웹제공"), code=code,
            content=co.cell(r, cc["content"]).value or "",
            connection=co.cell(r, cc["connection"]).value or "")
    out = []
    for no in sorted(meta):
        m = meta[no]
        m["items"] = {k: m["items"][k] for k in SECTION_ORDER if k in m["items"]}
        out.append(m)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    filled = sum(1 for m in out for s in m["items"].values() if s["content"])
    print(f"{os.path.basename(json_path)} 생성: 유산 {len(out)}건, 본문 {filled}개")

def rows(ws):
    """헤더(1행) 기준으로 각 행을 dict로. 빈 행은 건너뜀."""
    keys = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
    out = []
    for r in range(2, ws.max_row+1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        if all(v in (None, "") for v in vals): continue
        out.append({k: v for k, v in zip(keys, vals) if k})
    return out

def convert_config(xlsx_path, json_path):
    """levels/badges/sections 시트 → config.json (레벨·뱃지·항목 설정)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    levels, badges, sections = [], [], []
    if "levels" in wb.sheetnames:
        for x in rows(wb["levels"]):
            levels.append(dict(level=int(x["level"]), min=int(x["min"]), name=str(x["name"]).strip(),
                               icon=str(x.get("icon") or "").strip()))
        levels.sort(key=lambda v: v["min"])
    if "badges" in wb.sheetnames:
        for x in rows(wb["badges"]):
            badges.append(dict(type=str(x["type"]).strip(), key=str(x["key"]).strip(),
                               emoji=str(x.get("emoji") or "").strip(), name=str(x["name"]).strip(),
                               desc=str(x.get("desc") or "").strip()))
    if "sections" in wb.sheetnames:
        for x in rows(wb["sections"]):
            sections.append(dict(order=int(x["order"]), key=str(x["key"]).strip(),
                                 sub=str(x.get("sub") or "").strip()))
        sections.sort(key=lambda v: v["order"])
    config = dict(levels=levels, badges=badges, sections=sections)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=1)
    print(f"{os.path.basename(json_path)} 생성: 레벨 {len(levels)}단계, 뱃지 {len(badges)}종, 항목 {len(sections)}개")

if __name__ == "__main__":
    xlsx = find_xlsx()
    convert(xlsx, os.path.join(HERE, "content.json"))
    convert_config(xlsx, os.path.join(HERE, "config.json"))
